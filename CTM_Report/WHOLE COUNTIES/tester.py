#this is where i test functionality of snippets of code. once testing is done,
#the code gets moved over to the appropriate file/directory

import pandas as pd
from openpyxl import Workbook

wb = Workbook()
ws1 = wb.create_sheet("hello")
ws2 = wb.create_sheet("world")
wb.save("tester.xlsx")
    

"""
for row, index in data.itterows():
    print(data_fipsrow)
"""


    